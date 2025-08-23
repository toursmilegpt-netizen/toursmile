#!/usr/bin/env python3
"""
TourSmile Project Tracker Updater
Updates the Excel project tracker with current progress
"""
import openpyxl
from datetime import datetime
import sys

def update_project_tracker():
    try:
        # Load the existing tracker
        workbook = openpyxl.load_workbook('/app/TourSmile_Project_Tracker.xlsx')
        
        print("📊 Examining Excel file structure:")
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            
            # Print first few rows to understand structure
            for row_num in range(1, min(21, sheet.max_row + 1)):
                row_data = []
                for col_num in range(1, min(11, sheet.max_column + 1)):
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = cell.value
                    if value:
                        row_data.append(f"Col{col_num}: {value}")
                
                if row_data:
                    print(f"Row {row_num}: {' | '.join(row_data)}")
        
        return True
        
    except Exception as e:
        print(f"Error examining tracker: {e}")
        return False

if __name__ == "__main__":
    update_project_tracker()